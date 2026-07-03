#!/usr/bin/env python3
"""Fill 李涛绩效.xlsx completion status from Sandwich project work."""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment

XLSX = Path("/workspace/李涛绩效.xlsx")

# Column O = 15 完成情况
COMPLETION = {
    4: (
        "上半年持续跟进连续氢化建设项目，协调 CFCT、IEPE 与 SW 各方沟通；"
        "配合推进 HAZOP 后阶段布局设计、工程接口及跨团队信息对齐，保障项目按节点推进。"
    ),
    5: (
        "① 组织并参与英国侧可行性研究讲解及多轮技术对接，推动制备液相色谱+冻干机改造方案、"
        "布局及投资口径在 SW 与工程设备部之间达成一致；\n"
        "② 完成改造可行性研究消化与投资结构梳理（含直接/间接费用及预备费），"
        "形成多版组合汇报材料（HTML 简报），支撑管理层决策沟通；\n"
        "③ 协调 Hanbon 等设备供货信息、方案变更（DAC300 输送单元、废液罐、冻干隔离器等）"
        "及国内 TJ 团队经验输入；\n"
        "④ 推动 C1 模块五级密闭升级作为独立工作流纳入组合投资与周期统筹，"
        "明确与制备液相色谱改造同步交付要求；\n"
        "⑤ 对接 Clare/Woodley Cole 等高活实验室隔离器公用工程需求，"
        "组织国内工程与实验室专题会，输出中英对照答复材料。\n"
        "目前项目已形成可研级投资与进度基线，为下半年 FEED 及长周期设备采购决策提供支撑。"
    ),
    6: (
        "① 全程参与 Scitech 902 东侧扩建可行性研究（RIBA 第一阶段）讲解与纪要整理，"
        "输出详细中文汇总及对外沟通草稿，推动国内团队理解方案 Option 1 及关键假设；\n"
        "② 协助完成投资量级梳理（可行性阶段 OOM）及投资结构、风险与进度解读，"
        "编制 COO/CFO 管理层 HTML/PPT 汇报材料；\n"
        "③ 协调工程设备部与 SW 在拆除加氢厂房、首层净高、长周期设备采购分工等"
        "关键议题上的信息对齐与跟进；\n"
        "④ 配合凯总研发述职等场合，整理英国站点产能建设（扩建+改造+高活实验室）一页式材料。\n"
        "目前扩建可研已完成，概念设计（RIBA 2）按计划推进，为下半年方案深化与立项决策奠定基础。"
    ),
    7: (
        "按要求参与 Sandwich 站点年度预算相关沟通与数据支持；"
        "协调机器学习研讨小组相关技术交流安排（按上半年实际开展情况填报）。"
    ),
    8: (
        "配合生产管理相关重点工作推进，按部门安排完成电算化等专项支持任务。"
    ),
}

# Optional: refresh mid-year progress in 年度目标 (col G) for the three Sandwich sub-items
ANNUAL_TARGET_UPDATE = {
    4: (
        "1）连续氢化建设项目管理支持：\n"
        "目前项目已推进至 HAZOP 完成阶段。2026年将持续支持并协调 CFCT、IEPE 与 SW 各方，"
        "推进布局设计、工程采购、CE认证、测试验证等工作，做好跨团队沟通与协同，保障项目按计划整体推进。"
    ),
    5: (
        "2）HPLC和冻干机建设项目管理支持：\n"
        "目前项目已完成可行性研究（P01）及投资估算梳理，并形成组合汇报材料；"
        "2026年下半年将持续支持并协调工程设备部与 SW 方，推进 FEED 启动、长周期设备采购、"
        "C1 五级密闭升级协同及高活实验室接口需求澄清，保障项目有序实施。"
    ),
    6: (
        "3）902贴建项目管理支持：\n"
        "目前项目已完成东侧扩建可行性研究（RIBA 1，2026年5月）及管理层汇报材料编制；"
        "2026年下半年将持续支持并协调工程设备部与 SW 方，推进 RIBA 2 概念设计、"
        "关键方案决策（加氢厂房处置、设备采购分工等）及与国内团队的信息闭环，确保项目按计划开展。"
    ),
}


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["李涛自评"]
    wrap = Alignment(wrap_text=True, vertical="top")

    for row, text in COMPLETION.items():
        cell = ws.cell(row=row, column=15)
        cell.value = text
        cell.alignment = wrap

    for row, text in ANNUAL_TARGET_UPDATE.items():
        cell = ws.cell(row=row, column=7)
        cell.value = text
        cell.alignment = wrap

    wb.save(XLSX)
    print(f"Updated {XLSX}")


if __name__ == "__main__":
    main()
